# app/services/exporter.py
# ─────────────────────────────────────────────────────────────────────────────
# Builds a formatted .xlsx file from a user's job applications.
#
# WHY in-memory:
#   We never write to disk — we build the file in a BytesIO buffer and
#   stream it directly to the client. Faster, no cleanup needed, works
#   on Railway where the filesystem is ephemeral.
# ─────────────────────────────────────────────────────────────────────────────

import io
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import JobApplication


# ── Column definitions ─────────────────────────────────────────────────────
# Each tuple: (header label, model attribute or callable, column width)
COLUMNS: list[tuple] = [
    ("Company",         "company_name",    25),
    ("Role",            "role",            30),
    ("Company Type",    "company_type",    15),
    ("Status",          "status",          15),
    ("Source",          "source",          15),
    ("Applied Date",    "applied_date",    18),
    ("Location",        "location",        20),
    ("Remote",          "is_remote",       10),
    ("Match Score",     "match_score",     14),
    ("ATS Score",       "ats_score",       12),
    ("Salary Target",   "salary_target",   16),
    ("Market Min",      "salary_market_min", 14),
    ("Market Max",      "salary_market_max", 14),
    ("Currency",        "salary_currency", 10),
    ("Interview Date",  "interview_date",  18),
    ("Notes",           "notes",           40),
    ("Applied By",      "applied_by",      12),
    ("Job URL",         "job_url",         40),
]

# ── Status colours (background fill for the Status column) ────────────────
STATUS_COLORS: dict[str, str] = {
    "Applied":   "4F86C6",   # Blue
    "Interview": "F0A500",   # Amber
    "Offered":   "27AE60",   # Green
    "Rejected":  "E74C3C",   # Red
    "Ghosted":   "95A5A6",   # Grey
}


def build_xlsx(applications: list[JobApplication]) -> io.BytesIO:
    """
    Build a formatted .xlsx workbook from a list of JobApplication rows.

    Args:
        applications: List of JobApplication model instances.

    Returns:
        io.BytesIO: In-memory buffer containing the .xlsx file,
                    seeked back to position 0 ready for send_file().
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # ── Header row styles ──────────────────────────────────────────────────
    header_font    = Font(bold=True, color="FFFFFF", size=11)
    header_fill    = PatternFill("solid", fgColor="1E293B")   # Dark slate
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        bottom=Side(style="thin", color="334155"),
        right=Side(style="thin", color="334155"),
    )

    # ── Write headers ──────────────────────────────────────────────────────
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

        # Set column width
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze the header row so it stays visible when scrolling
    ws.freeze_panes = "A2"

    # ── Write data rows ────────────────────────────────────────────────────
    for row_idx, app in enumerate(applications, start=2):
        is_even: bool = row_idx % 2 == 0
        row_fill = PatternFill("solid", fgColor="F8FAFC" if is_even else "FFFFFF")

        for col_idx, (_, attr, _) in enumerate(COLUMNS, start=1):
            value = _get_value(app, attr)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=(attr == "notes"))
            cell.border    = thin_border

            # Colour the Status cell based on its value
            if attr == "status" and value in STATUS_COLORS:
                cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[value])
                cell.font = Font(bold=True, color="FFFFFF")
            else:
                cell.fill = row_fill

    # ── Summary row ────────────────────────────────────────────────────────
    # Add a totals row at the bottom with application counts per status
    if applications:
        _add_summary_sheet(wb, applications)

    # ── Auto-filter (dropdown arrows on headers) ───────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # ── Write to buffer ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)   # Rewind so send_file() reads from the beginning

    return buffer


def _get_value(app: JobApplication, attr: str):
    """
    Extract a display-ready value from a JobApplication for a given attribute.

    Args:
        app:  The JobApplication instance.
        attr: The attribute name from COLUMNS.

    Returns:
        A plain Python value safe to write into an Excel cell.
    """
    value = getattr(app, attr, None)

    if value is None:
        return ""

    # Format datetime objects as readable strings
    if isinstance(value, datetime):
        return value.strftime("%d %b %Y")

    # Convert booleans to readable text
    if isinstance(value, bool):
        return "Yes" if value else "No"

    # Round float scores
    if isinstance(value, float):
        return round(value, 2)

    return value


def _add_summary_sheet(wb: openpyxl.Workbook, apps: list[JobApplication]) -> None:
    """
    Add a second sheet with a status summary table.

    Args:
        wb:   The workbook to add the sheet to.
        apps: All applications (used to compute counts).
    """
    ws = wb.create_sheet(title="Summary")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E293B")

    # Headers
    ws["A1"] = "Status"
    ws["B1"] = "Count"
    ws["C1"] = "Percentage"

    for cell in [ws["A1"], ws["B1"], ws["C1"]]:
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14

    total: int = len(apps)

    # Count per status
    for row, status in enumerate(["Applied", "Interview", "Offered",
                                   "Rejected", "Ghosted"], start=2):
        count: int = sum(1 for a in apps if a.status == status)
        pct: str   = f"{round(count / total * 100, 1)}%" if total else "0%"

        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=3, value=pct)

        # Colour each status row
        if status in STATUS_COLORS:
            fill = PatternFill("solid", fgColor=STATUS_COLORS[status])
            ws.cell(row=row, column=1).fill = fill
            ws.cell(row=row, column=1).font = Font(bold=True, color="FFFFFF")

    # Total row
    total_row: int = len(["Applied", "Interview", "Offered", "Rejected", "Ghosted"]) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=total).font   = Font(bold=True)
    ws.cell(row=total_row, column=3, value="100%").font  = Font(bold=True)